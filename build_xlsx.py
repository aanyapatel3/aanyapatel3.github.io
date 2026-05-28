"""Build StockAnalysis.xlsx from the equity-research note.

Each model table is its own sheet. The final column of data sheets is a
clickable hyperlink to the public source the figure was drawn from. A master
"Sources" sheet lists every link. Figures marked "E" are analyst estimates.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ---- Source registry: key -> (display text, url) ----
SRC = {
    "amgn_guide":   ("Seeking Alpha - Amgen 2026 guidance", "https://seekingalpha.com/news/4546762-amgen-outlines-37b-38_4b-2026-revenue-target-as-repatha-evenity-and-tezspire-drive-momentum"),
    "amgn_stats":   ("stockanalysis.com - AMGN statistics", "https://stockanalysis.com/stocks/amgn/statistics/"),
    "amgn_q1_8k":   ("SEC - Amgen Q1 2026 8-K", "https://www.sec.gov/Archives/edgar/data/0000318154/000031815426000054/amgn-20260331earningsrelea.htm"),
    "amgn_q1_news": ("Amgen newsroom - Q1 2026 results", "https://www.amgen.com/newsroom/press-releases/2026/04/amgen-reports-first-quarter-2026-financial-results"),
    "amgn_fy25_8k": ("SEC - Amgen FY2025 8-K (margin/tax)", "https://www.sec.gov/Archives/edgar/data/0000318154/000031815426000003/amgn-20251231earningsrelea.htm"),
    "maritide_ph2": ("Amgen - MariTide Phase 2 weight loss", "https://investors.amgen.com/news-releases/news-release-details/amgen-announces-robust-weight-loss-maritide-people-living/"),
    "pipeline":     ("Labiotech - Amgen pipeline strategy 2026", "https://www.labiotech.eu/in-depth/amgen-pipeline-strategy/"),
    "maritide_pk":  ("FiercePharma/Evaluate - obesity peak sales", "https://www.fiercepharma.com/pharma/2030-eli-lilly-will-generate-113b-drug-sales-including-62b-mounjaro-zepbound-evaluate"),
    "dhr_seg":      ("Investing.com - Danaher Q4 2025 segments", "https://www.investing.com/news/company-news/danaher-q4-2025-slides-revenue-growth-solid-despite-margin-pressure-93CH-4470615"),
    "dhr_overview": ("StockTitan - Danaher 2025 overview", "https://www.stocktitan.net/sec-filings/DHR/ars-danaher-corp-de-sec-filing-3f1740b11e9d.html"),
    "dhr_tikr":     ("TIKR - Danaher forecast/valuation", "https://www.tikr.com/blog/danaher-stock-is-down-19-in-the-past-3-months-could-dhr-reach-248"),
    "sdz_fy25":     ("Investing.com - Sandoz FY2025 results", "https://www.investing.com/news/company-news/sandoz-fy-2025-slides-biosimilar-surge-drives-24-stock-gain-93CH-4523555"),
    "sdz_growth":   ("Pharmaceutical Daily - Sandoz 2025/2026 outlook", "https://pharmaceuticaldaily.com/sandoz-reports-7-revenue-growth-in-2025-expands-biosimilars-and-projects-stronger-2026-outlook/"),
    "sdz_invest":   ("Investing.com - Sandoz price/market data", "https://www.investing.com/equities/sandoz"),
}

# ---- styling helpers ----
HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F3864")
NOTE_FONT = Font(italic=True, size=9, color="666666")
TOTAL_FONT = Font(bold=True)
LINK_FONT = Font(color="0563C1", underline="single", size=10)
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

wb = Workbook()


def link_cell(ws, cell, src_key):
    disp, url = SRC[src_key]
    cell.value = disp
    cell.hyperlink = url
    cell.font = LINK_FONT
    cell.alignment = LEFT


def add_sheet(name, title, note, headers, rows, widths=None, total_rows=()):
    """rows: list of lists; last element of each row is a source key (or None)."""
    ws = wb.create_sheet(name)
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    r = 2
    if note:
        ws.cell(r, 1, note).font = NOTE_FONT
        r += 1
    hdr_row = r + 1
    for c, h in enumerate(headers, 1):
        cell = ws.cell(hdr_row, c, h)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = CENTER
        cell.border = BORDER
    for i, row in enumerate(rows):
        rr = hdr_row + 1 + i
        src_key = row[-1]
        data = row[:-1]
        for c, val in enumerate(data, 1):
            cell = ws.cell(rr, c, val)
            cell.border = BORDER
            cell.alignment = CENTER if c > 1 else LEFT
            if i in total_rows:
                cell.font = TOTAL_FONT
        src_cell = ws.cell(rr, len(headers), "")
        src_cell.border = BORDER
        if src_key:
            link_cell(ws, src_cell, src_key)
    if widths:
        for col, w in widths.items():
            ws.column_dimensions[col].width = w
    ws.freeze_panes = ws.cell(hdr_row + 1, 1)
    return ws


# ===== 1. Summary =====
add_sheet(
    "Summary", "Equity Research Summary - ratings & targets (late May 2026)",
    "Educational exercise; NOT investment advice. Figures from public sources as of late May 2026. 'E' = illustrative analyst estimate.",
    ["Company", "Ticker", "Segment", "Price*", "Rating", "Base 12-mo target", "Implied return**", "Source"],
    [
        ["Amgen", "AMGN (Nasdaq)", "Big-cap biopharma (drugs)", "~$326", "BUY", "~$365", "~+12% + 2.8% yield", "amgn_stats"],
        ["Danaher", "DHR (NYSE)", "Picks & shovels for biopharma", "~$190", "BUY / Accumulate", "~$225", "~+18% + 0.5% yield", "dhr_tikr"],
        ["Sandoz", "SDZ (SIX, Swiss)", "Generics + biosimilars", "~CHF 66", "HOLD", "~CHF 70", "~+6% + 1.2% yield", "sdz_invest"],
    ],
    widths={"A": 12, "B": 16, "C": 30, "D": 10, "E": 16, "F": 18, "G": 22, "H": 38},
)

# ===== 2. Amgen revenue build =====
add_sheet(
    "Amgen-RevenueBuild", "Amgen - bottom-up revenue build ($M)",
    "FY2025A = scraped actual (source per row). 2026E/2027E columns are analyst estimates (E).",
    ["Product", "FY2025A", "'26E gr.", "2026E", "'27E gr.", "2027E", "Rationale", "Source"],
    [
        ["Prolia", 4410, "-30%", 3087, "-25%", 2315, "Biosimilar cliff (Q1'26 -34%)", "amgn_q1_news"],
        ["Repatha", 3020, "+28%", 3866, "+20%", 4639, "FY25 +36%; VESALIUS-CV expands patients", "amgn_fy25_8k"],
        ["EVENITY", 2100, "+25%", 2625, "+18%", 3098, "FY25 +34%; volume momentum", "amgn_fy25_8k"],
        ["TEZSPIRE", 1480, "+28%", 1894, "+22%", 2311, "FY25 +52%; asthma share gains", "amgn_fy25_8k"],
        ["TEPEZZA", 1900, "+8%", 2052, "+6%", 2175, "Normalizing after inventory swings", "amgn_fy25_8k"],
        ["BLINCYTO", 1560, "+15%", 1794, "+12%", 2009, "Q1'26 +12%; broad prescribing", "amgn_q1_news"],
        ["KRYSTEXXA", 1340, "+10%", 1474, "+8%", 1592, "Q1'26 +8%; price + volume", "amgn_q1_news"],
        ["Otezla", 2270, "-2%", 2225, "-2%", 2181, "Mature / competitive", "amgn_fy25_8k"],
        ["XGEVA", 2080, "-12%", 1830, "-12%", 1610, "Biosimilar pressure", "amgn_fy25_8k"],
        ["Other (see Other-Bucket sheet)", 14988, "-1%", 14838, "flat", 14838, "Net of legacy erosion vs new launches", "amgn_q1_8k"],
        ["TOTAL PRODUCT SALES", 35148, "", 35685, "", 36768, "Bottom-up sum", None],
        ["+ Other revenue", 1550, "", 1600, "", 1650, "Royalties etc.", None],
        ["= TOTAL REVENUE", 36698, "", 37285, "", 38418, "vs guidance $37.0-38.4bn (2026)", "amgn_guide"],
    ],
    widths={"A": 30, "B": 10, "C": 9, "D": 10, "E": 9, "F": 10, "G": 38, "H": 34},
    total_rows={10, 11, 12},
)

# ===== 3. Amgen Other bucket =====
add_sheet(
    "Amgen-OtherBucket", "Amgen - decomposition of the 'Other' bucket ($M)",
    "Cluster dollar splits are ESTIMATES triangulated from quarterly disclosures, not exact line items.",
    ["Cluster", "FY2025A", "'26E gr.", "2026E", "'27E gr.", "2027E", "Read", "Source"],
    [
        ["Enbrel", 2800, "-10%", 2520, "-10%", 2268, "Slow secular decline, no biosimilar yet", "amgn_q1_8k"],
        ["Kyprolis", 1500, "-5%", 1425, "-8%", 1311, "Generic/competitive pressure building", "amgn_q1_8k"],
        ["Established (EPOGEN/Aranesp/Neulasta/Parsabiv)", 2000, "-10%", 1800, "-10%", 1620, "Legacy run-off", "amgn_q1_8k"],
        ["Mature biosimilars (Amjevita/MVASI/Kanjinti/Riabni)", 1900, "-8%", 1748, "-8%", 1608, "Price competition", "amgn_q1_8k"],
        ["Vectibix", 1100, "+2%", 1122, "+2%", 1144, "Stable oncology", "amgn_q1_8k"],
        ["Nplate", 1450, "+5%", 1523, "+4%", 1584, "Steady grower", "amgn_q1_8k"],
        ["Newer launches (IMDELLTRA/UPLIZNA/LUMAKRAS/TAVNEOS/Aimovig/Pavblu/Wezlana)", 2700, "+28%", 3456, "+24%", 4285, "The engine - ramping hard", "amgn_q1_8k"],
        ["Residual / smaller products", 1538, "0%", 1538, "0%", 1538, "Catch-all", None],
        ["'OTHER' TOTAL", 14988, "+0.9%", 15130, "+1.5%", 15358, "Confirms ~flat assumption", None],
    ],
    widths={"A": 48, "B": 10, "C": 9, "D": 10, "E": 9, "F": 10, "G": 32, "H": 30},
    total_rows={8},
)

# ===== 4. Amgen DCF =====
ws = wb.create_sheet("Amgen-DCF")
ws["A1"] = "Amgen - unlevered FCFF DCF"
ws["A1"].font = TITLE_FONT
ws["A2"] = "Anchored to FY2025 actuals; FCFF figures are analyst estimates (E). WACC rounded up to 7.5% to haircut leverage/pipeline risk."
ws["A2"].font = NOTE_FONT
inputs = [
    ("INPUT", "VALUE", "SOURCE"),
    ("Total revenue FY2025 ($bn)", 36.8, "amgn_fy25_8k"),
    ("Non-GAAP operating income ($bn)", 16.2, "amgn_fy25_8k"),
    ("Non-GAAP operating margin", "46.1%", "amgn_fy25_8k"),
    ("Effective tax rate", "18%", "amgn_fy25_8k"),
    ("Net debt ($bn)", 46.3, "amgn_stats"),
    ("Shares outstanding (m)", 539.7, "amgn_stats"),
    ("Cost of equity (Rf 4.3% + B0.6 x ERP 5%)", "7.3%", None),
    ("After-tax cost of debt (5% x (1-18%))", "4.1%", None),
    ("WACC used", "7.5%", None),
    ("Terminal growth (g)", "2.5%", None),
]
r = 4
for row in inputs:
    label, val, src = row
    c1 = ws.cell(r, 1, label)
    c2 = ws.cell(r, 2, val)
    c3 = ws.cell(r, 3, "")
    for c in (c1, c2, c3):
        c.border = BORDER
    if r == 4:
        for c in (c1, c2, c3):
            c.fill = HDR_FILL
            c.font = HDR_FONT
            c.alignment = CENTER
    elif src:
        link_cell(ws, c3, src)
    r += 1
r += 1
ws.cell(r, 1, "FCFF PROJECTION & DISCOUNTING ($bn)").font = Font(bold=True, color="1F3864")
r += 1
fcff_hdr = ["Line", "2026E", "2027E", "2028E", "2029E", "2030E"]
for c, h in enumerate(fcff_hdr, 1):
    cell = ws.cell(r, c, h)
    cell.fill = HDR_FILL; cell.font = HDR_FONT; cell.alignment = CENTER; cell.border = BORDER
fcff_rows = [
    ["Unlevered FCFF", 11.5, 12.1, 12.8, 13.6, 14.3],
    ["Discount factor @7.5%", 0.930, 0.865, 0.805, 0.749, 0.697],
    ["PV of FCFF", 10.7, 10.5, 10.3, 10.2, 10.0],
]
r += 1
for row in fcff_rows:
    for c, v in enumerate(row, 1):
        cell = ws.cell(r, c, v); cell.border = BORDER; cell.alignment = CENTER if c > 1 else LEFT
    r += 1
r += 1
outputs = [
    ("Sum PV of explicit FCFF ($bn)", 51.6),
    ("Terminal value at 2030 ($bn)", 293.0),
    ("PV of terminal value ($bn)", 204.0),
    ("Enterprise value ($bn)", 256.0),
    ("Less: net debt ($bn)", -46.3),
    ("Equity value ($bn)", 210.0),
    ("Implied value per share ($)", 388.0),
]
for label, val in outputs:
    c1 = ws.cell(r, 1, label); c2 = ws.cell(r, 2, val)
    c1.font = TOTAL_FONT
    if "per share" in label:
        c2.font = Font(bold=True, color="C00000", size=12)
    for c in (c1, c2):
        c.border = BORDER
    r += 1
ws.column_dimensions["A"].width = 42
for col in ["B", "C", "D", "E", "F"]:
    ws.column_dimensions[col].width = 12
ws.column_dimensions["C"].width = 38

# ===== 5. Amgen DCF sensitivity =====
add_sheet(
    "Amgen-DCF-Sensitivity", "Amgen DCF - implied $/share sensitivity",
    "Rows = WACC, columns = terminal growth (g). Base case shaded conceptually = 7.5% / 2.5% = $388.",
    ["WACC \\ g", "2.0%", "2.5%", "3.0%", "Source"],
    [
        ["7.0%", 396, 442, 498, None],
        ["7.5%", 352, 388, 432, None],
        ["8.0%", 315, 345, 380, None],
    ],
    widths={"A": 12, "B": 10, "C": 10, "D": 10, "E": 20},
)

# ===== 6. Amgen valuation scenarios =====
add_sheet(
    "Amgen-Valuation", "Amgen - P/E valuation scenarios",
    "Cross-checks the DCF. EPS within / around FY2026 non-GAAP guide of $21.60-23.00.",
    ["Scenario", "EPS used", "Multiple", "Implied price", "vs ~$326", "Source"],
    [
        ["Bear (MariTide miss, tax loss)", "$21.00", "12x", "$252", "-23%", "amgn_guide"],
        ["Base", "$22.50", "16x", "$360-365", "+11%", "amgn_guide"],
        ["Bull (obesity works)", "$23.50", "18x", "$423", "+30%", "amgn_guide"],
    ],
    widths={"A": 32, "B": 12, "C": 10, "D": 14, "E": 10, "F": 34},
)

# ===== 7. Amgen pipeline =====
add_sheet(
    "Amgen-Pipeline", "Amgen - risk-adjusted pipeline ($bn peak sales)",
    "PoS and peak-sales are analyst estimates. MariTide peak range: Evaluate ~$3.7bn to UBS $10bn+.",
    ["Asset", "Mechanism / indication", "Status", "Est. launch", "Unadj. peak", "PoS", "Risk-adj. peak", "In base?", "Source"],
    [
        ["MariTide", "GLP-1 agonist / GIPR antagonist, monthly dosing - obesity (+CV/HF/OSA)", "Ph3 (6 MARITIME studies)", "~2027-28", 6.0, "60%", 3.6, "No - upside", "maritide_ph2"],
        ["Olpasiran", "Lp(a) siRNA - CV risk", "Ph3 OCEAN(a)", "~2027-28", 2.5, "50%", 1.25, "No - upside", "pipeline"],
        ["Bemarituzumab", "First-in-class FGFR2b mAb - gastric cancer", "Ph3", "~2027", 1.5, "55%", 0.8, "No - upside", "pipeline"],
        ["IMDELLTRA", "DLL3 BiTE - small-cell lung (2L approved, 1L ongoing)", "Approved + expansion", "launched", 3.0, "80%", 2.4, "Partly", "pipeline"],
        ["UPLIZNA", "anti-CD19 - new IgG4-RD & gMG approvals", "Approved/expanding", "launched", 2.0, "85%", 1.7, "Partly", "pipeline"],
        ["Repatha primary-prevention", "PCSK9 - VESALIUS-CV positive", "Label expansion", "2026-27", "-", "-", "-", "Yes", "amgn_q1_8k"],
        ["MariTide peak-sales range (external)", "Evaluate $3.7bn (low) to UBS $10bn+ (high)", "-", "-", "-", "-", "-", "-", "maritide_pk"],
    ],
    widths={"A": 26, "B": 44, "C": 22, "D": 11, "E": 11, "F": 7, "G": 12, "H": 11, "I": 32},
)

# ===== 8. Danaher segments =====
add_sheet(
    "Danaher-Segments", "Danaher - segment revenue build ($bn)",
    "FY2025 actuals + margins per source. 2026E/2027E and Masimo phasing are analyst estimates (E).",
    ["Segment", "FY2025 rev", "Adj op margin", "'26E gr.", "2026E", "'27E gr.", "2027E", "Driver", "Source"],
    [
        ["Biotechnology (bioprocessing)", 7.3, "~38%", "+6%", 7.74, "+8%", 8.36, "Orders +30% - recovery engine", "dhr_seg"],
        ["Diagnostics (Cepheid, Beckman)", 10.0, "~31%", "+4%", 10.40, "+4%", 10.82, "Cepheid respiratory; China headwind", "dhr_seg"],
        ["Life Sciences (instruments/tools)", 7.3, "~21%", "+2%", 7.45, "+3%", 7.67, "Sluggish academic/biotech capex", "dhr_seg"],
        ["ORGANIC TOTAL", 24.6, "~28%", "+4%", 25.6, "+5%", 26.9, "In line with 3-6% core guide", "dhr_overview"],
        ["+ Masimo (patient monitoring)", "-", "-", "-", 1.0, "-", 2.2, "Closes 2H26, accretive yr 1", "dhr_tikr"],
        ["TOTAL REVENUE", 24.6, "", "", 26.6, "", 29.0, "", None],
    ],
    widths={"A": 32, "B": 11, "C": 13, "D": 9, "E": 9, "F": 9, "G": 9, "H": 32, "I": 30},
    total_rows={3, 5},
)

# ===== 9. Danaher valuation =====
add_sheet(
    "Danaher-Valuation", "Danaher - P/E valuation scenarios",
    "FY2026 adj EPS guide $8.35-8.55; 2027E ~$9.30 is analyst estimate.",
    ["Scenario", "EPS used", "Multiple", "Implied price", "vs ~$190", "Source"],
    [
        ["Bear (recovery stalls)", "2026E $8.45", "20x", "$169", "-11%", "dhr_tikr"],
        ["Base", "2027E ~$9.30", "24x", "$223", "+17%", "dhr_tikr"],
        ["Bull (full upcycle)", "2027E ~$9.60", "28x", "$269", "+42%", "dhr_tikr"],
    ],
    widths={"A": 26, "B": 14, "C": 10, "D": 14, "E": 10, "F": 30},
)

# ===== 10. Sandoz build =====
add_sheet(
    "Sandoz-Build", "Sandoz - revenue build: generics vs biosimilars ($bn)",
    "FY2025 actuals per source. 2026E/2027E are analyst estimates (E). Reports in USD; listed in CHF.",
    ["Segment", "FY2025", "'26E gr.", "2026E", "'27E gr.", "2027E", "Notes", "Source"],
    [
        ["Biosimilars", 3.29, "+14%", 3.75, "+15%", 4.31, "Hyrimoz, Pyzchiva, Omnitrope, denosumab", "sdz_fy25"],
        ["Generics", 7.79, "-1%", 7.71, "flat", 7.71, "Pricing -low/mid SD; portfolio pruning", "sdz_fy25"],
        ["TOTAL NET SALES", 11.09, "+5%", 11.46, "+5%", 12.02, "Mid-single-digit, in line w/ guide", "sdz_growth"],
        ["Biosimilar mix %", "30%", "", "~33%", "", "~36%", "The margin-expansion lever", None],
    ],
    widths={"A": 20, "B": 10, "C": 9, "D": 10, "E": 9, "F": 10, "G": 36, "H": 34},
    total_rows={2},
)

# ===== 11. Sandoz valuation =====
add_sheet(
    "Sandoz-Valuation", "Sandoz - EV/EBITDA valuation scenarios",
    "On ~$2.7bn 2026E core EBITDA. Net debt ~$3.6bn. Base case sits BELOW current ~CHF66 -> Hold.",
    ["Scenario", "2026E core EBITDA", "Multiple", "Implied equity (less $3.6bn debt)", "Rough price", "Source"],
    [
        ["Bear (pricing worse)", "$2.6bn", "9x", "~$19.8bn", "~CHF 37", "sdz_growth"],
        ["Base", "$2.7bn", "11x", "~$26.1bn", "~CHF 48-50", "sdz_growth"],
        ["Bull (margin + biosimilar beats)", "$2.9bn", "13x", "~$34bn", "~CHF 63", "sdz_growth"],
    ],
    widths={"A": 30, "B": 16, "C": 10, "D": 30, "E": 14, "F": 30},
)

# ===== 12. Sources master =====
ws = wb.create_sheet("Sources")
ws["A1"] = "Sources - all links used in this model"
ws["A1"].font = TITLE_FONT
ws["A2"] = "Click any link to open the source. Note: SEC and some press-release pages may block automated fetch but are viewable in a browser."
ws["A2"].font = NOTE_FONT
hdr = ["#", "Source", "Link"]
for c, h in enumerate(hdr, 1):
    cell = ws.cell(4, c, h); cell.fill = HDR_FILL; cell.font = HDR_FONT; cell.alignment = CENTER; cell.border = BORDER
for i, (key, (disp, url)) in enumerate(SRC.items(), 1):
    rr = 4 + i
    ws.cell(rr, 1, i).border = BORDER
    c2 = ws.cell(rr, 2, disp); c2.border = BORDER; c2.alignment = LEFT
    c3 = ws.cell(rr, 3, url); c3.hyperlink = url; c3.font = LINK_FONT; c3.border = BORDER; c3.alignment = LEFT
ws.column_dimensions["A"].width = 5
ws.column_dimensions["B"].width = 44
ws.column_dimensions["C"].width = 95

# remove default sheet & order
if "Sheet" in wb.sheetnames:
    del wb["Sheet"]

wb.save("/home/user/aanyapatel3.github.io/StockAnalysis.xlsx")
print("Wrote StockAnalysis.xlsx with sheets:", wb.sheetnames)
