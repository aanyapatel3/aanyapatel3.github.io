"""Build EquityResearch_Master.xlsx - ONE comprehensive workbook.

Covers all four names (Amgen, Danaher, Sandoz, Guardant Health) plus a plain-
English Read-Me and a Glossary. Every model table has: (1) an intro line saying
what the table is, (2) a Source column with clickable hyperlinks, and (3) a
"What this means" takeaway in plain English. Figures marked "E" are estimates.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------- Source registry (display text, url) ----------
SRC = {
    # Amgen / Danaher / Sandoz
    "amgn_guide":   ("Seeking Alpha - Amgen 2026 guidance", "https://seekingalpha.com/news/4546762-amgen-outlines-37b-38_4b-2026-revenue-target-as-repatha-evenity-and-tezspire-drive-momentum"),
    "amgn_stats":   ("stockanalysis.com - AMGN statistics", "https://stockanalysis.com/stocks/amgn/statistics/"),
    "amgn_q1_8k":   ("SEC - Amgen Q1 2026 8-K", "https://www.sec.gov/Archives/edgar/data/0000318154/000031815426000054/amgn-20260331earningsrelea.htm"),
    "amgn_q1_news": ("Amgen newsroom - Q1 2026 results", "https://www.amgen.com/newsroom/press-releases/2026/04/amgen-reports-first-quarter-2026-financial-results"),
    "amgn_fy25_8k": ("SEC - Amgen FY2025 8-K (margin/tax)", "https://www.sec.gov/Archives/edgar/data/0000318154/000031815426000003/amgn-20251231earningsrelea.htm"),
    "maritide_ph2": ("Amgen - MariTide Phase 2 weight loss", "https://investors.amgen.com/news-releases/news-release-details/amgen-announces-robust-weight-loss-maritide-people-living/"),
    "pipeline":     ("Labiotech - Amgen pipeline strategy 2026", "https://www.labiotech.eu/in-depth/amgen-pipeline-strategy/"),
    "maritide_pk":  ("FiercePharma/Evaluate - obesity peak sales", "https://www.fiercepharma.com/pharma/2030-eli-lilly-will-generate-113b-drug-sales-including-62b-mounjaro-zepbound-evaluate"),
    "dhr_seg":      ("Investing.com - Danaher Q4 2025 segments", "https://www.investing.com/news/company-news/danaher-q4-2025-slides-revenue-growth-solid-despite-margin-pressure-93CH-4470615"),
    "dhr_overview": ("StockTitan - Danaher 2025 overview", "https://www.stocktitan.net/sec-filings/DHR/ars-danaher-corp-de-sec-filing-3f1740b11e9d.html"),
    "dhr_tikr":     ("TIKR - Danaher forecast/valuation", "https://www.tikr.com/blog/danaher-stock-is-down-19-in-the-past-3-months-could-dhr-reach-248"),
    "sdz_fy25":     ("Investing.com - Sandoz FY2025 results", "https://www.investing.com/news/company-news/sandoz-fy-2025-slides-biosimilar-surge-drives-24-stock-gain-93CH-4523555"),
    "sdz_growth":   ("Pharmaceutical Daily - Sandoz 2025/2026 outlook", "https://pharmaceuticaldaily.com/sandoz-reports-7-revenue-growth-in-2025-expands-biosimilars-and-projects-stronger-2026-outlook/"),
    "sdz_invest":   ("Investing.com - Sandoz price/market data", "https://www.investing.com/equities/sandoz"),
    # Guardant Health
    "gh_mktcap":    ("stockanalysis.com - GH market cap & price", "https://stockanalysis.com/stocks/gh/market-cap/"),
    "gh_overview":  ("WallStreetZen - GH overview", "https://www.wallstreetzen.com/stocks/us/nasdaq/gh"),
    "gh_q1_tr":     ("Motley Fool - GH Q1 2026 transcript", "https://www.fool.com/earnings/call-transcripts/2026/05/08/guardant-health-gh-q1-2026-earnings-transcript/"),
    "gh_guidance":  ("StockTitan - GH boosts 2026 guidance (Q1 +48%)", "https://www.stocktitan.net/sec-filings/GH/8-k-guardant-health-inc-reports-material-event-e2c9da59af8b.html"),
    "gh_fy25":      ("BioSpace - GH Q4 & FY2025 results", "https://www.biospace.com/press-releases/guardant-health-reports-fourth-quarter-and-full-year-2025-financial-results-and-provides-2026-outlook"),
    "gh_q1_bw":     ("Business Wire - GH Q1 2026 results", "https://www.businesswire.com/news/home/20260507017149/en/Guardant-Health-Reports-First-Quarter-2026-Financial-Results-and-Increases-2026-Revenue-Guidance"),
    "shield_adlt":  ("Business Wire - Shield ADLT $1,495 Medicare rate", "https://www.businesswire.com/news/home/20250311660519/en/Guardant-Health-Receives-ADLT-Status-From-CMS-for-Shield-Blood-Test"),
    "shield_fda":   ("AdvaMed - Shield FDA approval (primary CRC screen)", "https://www.advamed.org/industry-updates/news/guardant-healths-shield-blood-test-approved-by-fda-as-a-primary-screening-option-clearing-path-for-medicare-reimbursement-and-a-new-era-of-colorectal-cancer-screening/"),
    "shield_acs":   ("GH IR - ACS recommends Shield in CRC guidelines", "https://investors.guardanthealth.com/press-releases/press-releases/2026/American-Cancer-Society-Recommends-Guardant-Healths-Shield-Blood-Test-in-Updated-Colorectal-Cancer-Screening-Guidelines/default.aspx"),
    "shield_cnn":   ("CNN - blood testing added to CRC screening recs", "https://www.cnn.com/2026/05/27/health/blood-test-colorectal-cancer-screening-wellness"),
    "shield_eclipse":("OncLive - Shield ECLIPSE sensitivity/specificity", "https://www.onclive.com/view/shield-blood-test-yields-high-sensitivity-specificity-for-crc-screening-in-expansion-cohort-of-average-risk-adults"),
    "shield_cov":   ("shieldcancerscreen.com - coverage & support", "https://shieldcancerscreen.com/hcp/coverage-and-support/"),
    "exact_free":   ("MedCity News - Exact Sciences + Freenome blood test", "https://medcitynews.com/2025/08/exact-sciences-colorectal-cancer-screening-freenome-blood-test-liquid-biopsy-exas/"),
    "exact_guide":  ("Seeking Alpha - Exact Sciences 2025 guide $3.235B", "https://seekingalpha.com/news/4513917-exact-sciences-raises-2025-revenue-guidance-to-3_235b-amid-cologuard-plus-expansion-and"),
    "natera_10k":   ("StockTitan - Natera 2025 results ($2.3B)", "https://www.stocktitan.net/sec-filings/NTRA/10-k-natera-inc-files-annual-report-9c689b4cdac7.html"),
    "mrd_context":  ("Daily Upside - MRD market / Natera context", "https://www.thedailyupside.com/industries/healthcare/nateras-stronghold-on-cancer-recurrence-tests-lures-top-investoraustin-based-biotechnology-company-natera-was-valued-at-roughly-28-4-billion-as-of-thursday-more-even-than-the-21-5-billion-market-ca/"),
    "gh_forecast":  ("stockanalysis.com - GH analyst targets", "https://stockanalysis.com/stocks/gh/forecast/"),
    "btig":         ("Investing.com - BTIG raises GH target to $155", "https://www.investing.com/news/analyst-ratings/btig-raises-guardant-health-stock-price-target-on-fda-approval-93CH-4709281"),
    "gh_mb":        ("MarketBeat - GH forecast & targets", "https://www.marketbeat.com/stocks/NASDAQ/GH/forecast/"),
    "gh_mult":      ("Multiples.vc - GH valuation / EV multiples", "https://multiples.vc/public-comps/guardant-health-valuation-multiples"),
    "gh_raise":     ("Panabee - GH ~$717m equity + 0% converts", "https://www.panabee.com/news/guardant-health-secures-717-million-signaling-aggressive-debt-maturity-strategy"),
    "gh_burn":      ("AInvest - GH growth vs losses / cash burn", "https://www.ainvest.com/aime/share/guardant-healths-strong-revenue-growth-persistent-losses-sustainable-expansion-strategy-cash-burn-9d3949/"),
    "gh_10q":       ("SEC - GH Q1 2026 10-Q (shares)", "https://www.sec.gov/Archives/edgar/data/0001576280/000157628026000026/gh-20260331.htm"),
}

# ---------- styling ----------
NAVY = "1F3864"; LIGHT = "DCE6F1"; GREY = "F2F2F2"
HDR_FILL = PatternFill("solid", fgColor=NAVY)
SECTION_FILL = PatternFill("solid", fgColor=LIGHT)
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=15, color=NAVY)
INTRO_FONT = Font(italic=True, size=10, color="333333")
NOTE_FONT = Font(italic=True, size=10, color="375623")
TOTAL_FONT = Font(bold=True)
LINK_FONT = Font(color="0563C1", underline="single", size=10)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)

wb = Workbook()


def link_cell(ws, cell, src_key):
    if not src_key:
        return
    disp, url = SRC[src_key]
    cell.value = disp
    cell.hyperlink = url
    cell.font = LINK_FONT
    cell.alignment = LEFT


def add_sheet(name, title, intro, headers, rows, takeaway, widths=None, total_rows=()):
    ws = wb.create_sheet(name)
    ncols = len(headers)
    last_col = get_column_letter(ncols)
    # title
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws.merge_cells(f"A1:{last_col}1")
    # intro
    ws["A2"] = intro
    ws["A2"].font = INTRO_FONT
    ws["A2"].alignment = LEFT_TOP
    ws.merge_cells(f"A2:{last_col}2")
    ws.row_dimensions[2].height = 42
    # header
    hr = 4
    for c, h in enumerate(headers, 1):
        cell = ws.cell(hr, c, h)
        cell.fill = HDR_FILL; cell.font = HDR_FONT; cell.alignment = CENTER; cell.border = BORDER
    # data
    for i, row in enumerate(rows):
        rr = hr + 1 + i
        for c, val in enumerate(row[:-1], 1):
            cell = ws.cell(rr, c, val)
            cell.border = BORDER
            cell.alignment = CENTER if c > 1 else LEFT
            if i in total_rows:
                cell.font = TOTAL_FONT
                cell.fill = PatternFill("solid", fgColor=GREY)
        sc = ws.cell(rr, ncols, "")
        sc.border = BORDER
        link_cell(ws, sc, row[-1])
        if i in total_rows:
            sc.fill = PatternFill("solid", fgColor=GREY)
    # takeaway
    tr = hr + 1 + len(rows) + 1
    lbl = ws.cell(tr, 1, "What this means:")
    lbl.font = Font(bold=True, color="375623", size=10)
    ws.cell(tr + 1, 1, takeaway).font = NOTE_FONT
    ws.cell(tr + 1, 1).alignment = LEFT_TOP
    ws.merge_cells(f"A{tr+1}:{last_col}{tr+1}")
    ws.row_dimensions[tr + 1].height = max(30, 14 * (len(takeaway) // (12 * ncols) + 1))
    if widths:
        for col, w in widths.items():
            ws.column_dimensions[col].width = w
    ws.freeze_panes = ws.cell(hr + 1, 1)
    return ws


# ================= READ ME =================
ws = wb.active
ws.title = "READ ME"
ws["A1"] = "Equity Research - Master Workbook"
ws["A1"].font = Font(bold=True, size=18, color=NAVY)
readme = [
    "",
    "What this is: a junior-analyst study of four healthcare stocks, with the financial models, valuations and every source in one place.",
    "",
    "IMPORTANT: This is an EDUCATIONAL exercise, NOT investment advice. All figures are from public sources as of late May 2026 and are rounded.",
    "Anything marked 'E' is an illustrative analyst estimate (my own), not company-reported and not consensus.",
    "",
    "How to read each sheet:",
    "  - The italic line under each title explains what the table is.",
    "  - The dark-blue header row labels the columns.",
    "  - The last column 'Source' is a CLICKABLE link to where the number came from.",
    "  - A green 'What this means' note under each table explains the takeaway in plain English.",
    "",
    "The four companies and my ratings:",
    "  - Amgen (AMGN) ............ BUY .................. big-cap biopharma (sells drugs)",
    "  - Danaher (DHR) ........... BUY / Accumulate ..... 'picks & shovels' for biopharma (sells lab/bioprocessing equipment)",
    "  - Sandoz (SDZ) ............ HOLD ................. generics + biosimilars (low-cost copies of drugs)",
    "  - Guardant Health (GH) .... BUY (high risk) ...... liquid-biopsy cancer testing (blood tests for cancer)",
    "",
    "Tab guide:",
    "  Summary .............. all four ratings & price targets in one table",
    "  Glossary ............. plain-English definitions of every financial term used",
    "  Amgen-* .............. revenue build, 'Other' bucket, DCF, sensitivity, valuation, pipeline",
    "  Danaher-* ............ segment revenue build, valuation",
    "  Sandoz-* ............. generics-vs-biosimilars build, valuation",
    "  GH-* ................. financials, revenue build, Shield TAM, competition, valuation, catalysts",
    "  Sources .............. every link used, in one list",
]
for i, line in enumerate(readme, start=3):
    c = ws.cell(i, 1, line)
    c.font = Font(size=11, bold=line.endswith(":") or line.startswith("IMPORTANT"))
    c.alignment = LEFT
ws.column_dimensions["A"].width = 120

# ================= SUMMARY =================
add_sheet(
    "Summary", "Ratings & price targets - all four names (late May 2026)",
    "One-glance scorecard. 'Price' is the recent market price; 'Base target' is my 12-month fair-value estimate; 'Implied return' is the upside/downside to that target plus any dividend.",
    ["Company", "Ticker", "What it does", "Price*", "Rating", "Base 12-mo target", "Implied return**", "Source"],
    [
        ["Amgen", "AMGN (Nasdaq)", "Big-cap biopharma (sells drugs)", "~$326", "BUY", "~$365", "~+12% + 2.8% yield", "amgn_stats"],
        ["Danaher", "DHR (NYSE)", "Equipment/tools for biopharma", "~$190", "BUY / Accumulate", "~$225", "~+18% + 0.5% yield", "dhr_tikr"],
        ["Sandoz", "SDZ (SIX, Swiss)", "Generics + biosimilars", "~CHF 66", "HOLD", "~CHF 70", "~+6% + 1.2% yield", "sdz_invest"],
        ["Guardant Health", "GH (Nasdaq)", "Liquid-biopsy cancer testing", "~$127", "BUY (high risk)", "~$150", "~+18% (no dividend)", "gh_mktcap"],
    ],
    "Two 'buy the cheap compounder' ideas (Amgen, Danaher), one 'great business but fully priced' (Sandoz = Hold), and one high-risk/high-reward growth bet (Guardant). They all ride the same trend - more biologic medicine and more cancer testing - so they would tend to move together; a truly diversified portfolio would add a different sector.",
    widths={"A": 17, "B": 17, "C": 30, "D": 10, "E": 17, "F": 18, "G": 22, "H": 40},
)

# ================= GLOSSARY =================
gloss = [
    ("Revenue (sales)", "The total money a company brings in from selling its products/services before any costs.", None),
    ("Guidance", "Management's own forecast for next year's revenue or profit. Analysts check their model against it.", None),
    ("EPS (earnings per share)", "Net profit divided by the number of shares. The per-share profit you 'own'.", None),
    ("P/E (price-to-earnings)", "Share price divided by EPS. How many years of current profit you pay for the stock. Lower = cheaper.", None),
    ("Forward P/E", "Same idea but using NEXT year's expected EPS. We use this a lot here.", None),
    ("EV (enterprise value)", "Market cap + debt - cash. The total cost to buy the whole business including its debts.", None),
    ("EV/Sales", "Enterprise value divided by revenue. Used for fast-growing, not-yet-profitable firms (e.g., Guardant) where P/E doesn't work.", None),
    ("EBITDA", "Earnings Before Interest, Taxes, Depreciation & Amortization - a rough proxy for cash operating profit.", None),
    ("Core / adjusted EBITDA", "EBITDA with one-off items stripped out, to show the 'underlying' run-rate (Sandoz reports this).", None),
    ("Gross margin", "Revenue minus the direct cost of making the product, as a % of revenue. Higher = more profitable per sale.", None),
    ("Operating margin", "Profit after all operating costs, as a % of revenue.", None),
    ("Net debt", "Total debt minus cash. Positive = owes more than it holds; negative = net cash.", None),
    ("Free cash flow (FCF)", "Cash left after running the business and capital spending. Negative FCF ('burn') means it's spending more than it makes.", None),
    ("DCF", "Discounted Cash Flow - value a company by projecting its future cash and discounting it back to today's money.", None),
    ("FCFF", "Free Cash Flow to the Firm - the cash available to ALL investors (debt + equity), used in the DCF.", None),
    ("WACC", "Weighted Average Cost of Capital - the blended return investors require; the 'discount rate' in a DCF. Higher WACC = lower value.", None),
    ("Terminal growth (g)", "The assumed forever-growth rate after the forecast years, used to value the business into perpetuity.", None),
    ("Basis points (bps)", "Hundredths of a percent. 100 bps = 1.0%.", None),
    ("Biosimilar", "A near-copy of a biologic (large-molecule) drug, sold after the original loses patent - cheaper, lower-margin.", "sdz_fy25"),
    ("Generic", "A copy of a small-molecule drug after patent expiry. Very cheap, commoditized.", None),
    ("Patent cliff / erosion", "The steep sales drop when a drug loses patent protection and copies arrive (e.g., Amgen's Prolia).", "amgn_q1_news"),
    ("Pipeline", "A company's drugs still in development (not yet approved/selling). Future growth, but uncertain.", "pipeline"),
    ("PoS (prob. of success)", "Estimated % chance a pipeline drug gets approved. Used to 'risk-adjust' its potential sales.", None),
    ("Peak sales", "The maximum annual revenue a drug is expected to reach at full market uptake.", None),
    ("Bioprocessing", "The equipment/consumables used to manufacture biologic drugs - Danaher's key growth segment.", "dhr_seg"),
    ("Liquid biopsy", "A blood test that detects cancer DNA - Guardant's whole business. Less invasive than a tissue biopsy.", None),
    ("MRD", "Molecular Residual Disease - tiny traces of tumor DNA left after treatment; tested to catch recurrence early (Guardant Reveal vs Natera).", "natera_10k"),
    ("CRC", "Colorectal cancer - the cancer Guardant's Shield blood test screens for.", "shield_fda"),
    ("ASP", "Average Selling Price - the realized price per test/unit after discounts and payer mix.", None),
    ("ADLT / Medicare rate", "A special Medicare pricing status; Shield's locked $1,495 reimbursement runs through 2027.", "shield_adlt"),
    ("USPSTF", "US Preventive Services Task Force - its screening recommendation unlocks no-cost insurance coverage; the big GH catalyst.", "shield_fda"),
    ("TAM", "Total Addressable Market - the full revenue opportunity if you captured the whole market.", None),
    ("Bear / Base / Bull", "Three scenarios: pessimistic / most-likely / optimistic. Showing all three avoids a single fragile guess.", None),
]
add_sheet(
    "Glossary", "Glossary - what every term means (plain English)",
    "Every financial term used anywhere in this workbook, defined simply. Where a term is central to one company, the Source link points to a relevant reference.",
    ["Term", "Plain-English meaning", "Source"],
    [[t, d, s] for (t, d, s) in gloss],
    "If a sheet ever uses a word you don't recognize, it's defined here. The key ones for this workbook: P/E and DCF (how we value Amgen/Danaher), EV/Sales (how we value Guardant), and EBITDA (how we value Sandoz).",
    widths={"A": 26, "B": 96, "C": 40},
)

# ================= AMGEN =================
add_sheet(
    "Amgen-RevenueBuild", "Amgen - bottom-up revenue build ($ millions)",
    "We forecast each big drug separately (its own growth rate), then add them up and check the total against management's guidance. 'A' = actual, 'E' = estimate.",
    ["Product", "FY2025A", "'26E gr.", "2026E", "'27E gr.", "2027E", "Why (rationale)", "Source"],
    [
        ["Prolia", 4410, "-30%", 3087, "-25%", 2315, "Losing patent - cheap copies launched (Q1'26 -34%)", "amgn_q1_news"],
        ["Repatha", 3020, "+28%", 3866, "+20%", 4639, "Heart-disease drug; strong trial data widens use", "amgn_fy25_8k"],
        ["EVENITY", 2100, "+25%", 2625, "+18%", 3098, "Osteoporosis; volume momentum (+34% in '25)", "amgn_fy25_8k"],
        ["TEZSPIRE", 1480, "+28%", 1894, "+22%", 2311, "Asthma; fastest grower (+52% in '25)", "amgn_fy25_8k"],
        ["TEPEZZA", 1900, "+8%", 2052, "+6%", 2175, "Thyroid eye disease; normalizing", "amgn_fy25_8k"],
        ["BLINCYTO", 1560, "+15%", 1794, "+12%", 2009, "Blood cancer; broad uptake", "amgn_q1_news"],
        ["KRYSTEXXA", 1340, "+10%", 1474, "+8%", 1592, "Gout; price + volume", "amgn_q1_news"],
        ["Otezla", 2270, "-2%", 2225, "-2%", 2181, "Psoriasis; mature/competitive", "amgn_fy25_8k"],
        ["XGEVA", 2080, "-12%", 1830, "-12%", 1610, "Cancer bone support; biosimilar pressure", "amgn_fy25_8k"],
        ["Other (see Other-Bucket tab)", 14988, "~flat", 14838, "~flat", 14838, "~Half the company - decomposed on next tab", "amgn_q1_8k"],
        ["TOTAL PRODUCT SALES", 35148, "", 35685, "", 36768, "Sum of all drugs", None],
        ["+ Other revenue", 1550, "", 1600, "", 1650, "Royalties etc.", None],
        ["= TOTAL REVENUE", 36698, "", 37285, "", 38418, "Lands mid-guidance ($37.0-38.4bn for 2026)", "amgn_guide"],
    ],
    "The growers (Repatha, EVENITY, TEZSPIRE) are almost exactly cancelled out by the decline of Prolia (losing its patent). So the existing drug business is barely growing (~1-4%/yr) - which means the stock's upside really comes from the PIPELINE (see Amgen-Pipeline tab), not from today's products.",
    widths={"A": 30, "B": 10, "C": 9, "D": 10, "E": 9, "F": 10, "G": 40, "H": 34},
    total_rows={10, 11, 12},
)

add_sheet(
    "Amgen-OtherBucket", "Amgen - inside the 'Other' bucket ($ millions)",
    "The single 'Other' line above is ~$15bn (~43% of Amgen). Here we split it up to check it's really just flat, not hiding a problem. Cluster splits are estimates.",
    ["Cluster", "FY2025A", "'26E gr.", "2026E", "'27E gr.", "2027E", "What it is", "Source"],
    [
        ["Enbrel", 2800, "-10%", 2520, "-10%", 2268, "Old arthritis drug, slowly declining", "amgn_q1_8k"],
        ["Kyprolis", 1500, "-5%", 1425, "-8%", 1311, "Blood-cancer drug, competition rising", "amgn_q1_8k"],
        ["Established (EPOGEN/Aranesp/Neulasta/Parsabiv)", 2000, "-10%", 1800, "-10%", 1620, "Legacy products running off", "amgn_q1_8k"],
        ["Mature biosimilars (Amjevita/MVASI/Kanjinti/Riabni)", 1900, "-8%", 1748, "-8%", 1608, "Amgen's own copies, price competition", "amgn_q1_8k"],
        ["Vectibix", 1100, "+2%", 1122, "+2%", 1144, "Colorectal cancer; stable", "amgn_q1_8k"],
        ["Nplate", 1450, "+5%", 1523, "+4%", 1584, "Low-platelet disorder; steady grower", "amgn_q1_8k"],
        ["Newer launches (IMDELLTRA/UPLIZNA/LUMAKRAS/TAVNEOS/Aimovig/Pavblu/Wezlana)", 2700, "+28%", 3456, "+24%", 4285, "The young, fast-growing drugs - the engine", "amgn_q1_8k"],
        ["Residual / smaller products", 1538, "0%", 1538, "0%", 1538, "Everything else", None],
        ["'OTHER' TOTAL", 14988, "+0.9%", 15130, "+1.5%", 15358, "Net ~flat", None],
    ],
    "The ~flat 'Other' line is NOT lazy modeling - it's ~$8bn of old products shrinking ~9%/yr, almost exactly offset by newer launches growing ~25-28%. So the resilience of Amgen's boring half depends entirely on those new launches outrunning the old-drug decline.",
    widths={"A": 50, "B": 10, "C": 9, "D": 10, "E": 9, "F": 10, "G": 34, "H": 30},
    total_rows={8},
)

add_sheet(
    "Amgen-DCF", "Amgen - Discounted Cash Flow valuation ($ billions unless noted)",
    "A DCF values the company by its future cash. We project cash (FCFF), discount it to today using WACC, add it up, subtract debt, and divide by shares. See Glossary for any term.",
    ["Line item", "2026E", "2027E", "2028E", "2029E", "2030E", "Note", "Source"],
    [
        ["Unlevered free cash flow (FCFF)", 11.5, 12.1, 12.8, 13.6, 14.3, "Cash to all investors", "amgn_fy25_8k"],
        ["Discount factor @ 7.5% WACC", 0.930, 0.865, 0.805, 0.749, 0.697, "Today's-money conversion", None],
        ["Present value of FCFF", 10.7, 10.5, 10.3, 10.2, 10.0, "FCFF x discount factor", None],
        ["Sum of PV (years 2026-30)", 51.6, "", "", "", "", "Add the row above", None],
        ["+ PV of terminal value", 204.0, "", "", "", "", "Value of all years after 2030", None],
        ["= Enterprise value", 256.0, "", "", "", "", "Whole-firm value", None],
        ["- Net debt", -46.3, "", "", "", "", "Debt minus cash", "amgn_stats"],
        ["= Equity value", 210.0, "", "", "", "", "Value to shareholders", None],
        ["/ Shares (539.7m)", "", "", "", "", "", "", "amgn_stats"],
        ["= Value per share", 388.0, "", "", "", "", "DCF fair value (base)", None],
    ],
    "The DCF says Amgen is worth ~$388/share vs ~$326 today - i.e., the market is pricing it cheaply. Inputs: WACC (required return) 7.5%, terminal growth 2.5%, tax 18%. See the Sensitivity tab for how the answer changes if those assumptions move.",
    widths={"A": 32, "B": 10, "C": 10, "D": 10, "E": 10, "F": 10, "G": 26, "H": 30},
    total_rows={3, 5, 9},
)

add_sheet(
    "Amgen-DCF-Sensitivity", "Amgen - DCF sensitivity (implied $/share)",
    "A DCF answer depends on two key guesses: WACC (the discount rate) and g (forever-growth). This grid shows the value per share for each combination. Base case = 7.5% / 2.5% = $388.",
    ["WACC \\ growth (g)", "g = 2.0%", "g = 2.5%", "g = 3.0%", "Source"],
    [
        ["WACC 7.0%", 396, 442, 498, None],
        ["WACC 7.5%", 352, 388, 432, None],
        ["WACC 8.0%", 315, 345, 380, None],
    ],
    "Every cell except the most pessimistic corner ($315) is at or above today's ~$326 price. So even with conservative assumptions, the downside is small and the upside is large - that asymmetry is the core of the Buy.",
    widths={"A": 18, "B": 12, "C": 12, "D": 12, "E": 22},
)

add_sheet(
    "Amgen-Valuation", "Amgen - P/E valuation scenarios (cross-check on the DCF)",
    "A simpler valuation: take expected earnings per share (EPS) and multiply by a P/E multiple. Three scenarios. EPS is around the company's 2026 guidance of $21.60-23.00.",
    ["Scenario", "EPS used", "P/E multiple", "Implied price", "vs ~$326", "Source"],
    [
        ["Bear (pipeline miss, tax loss)", "$21.00", "12x", "$252", "-23%", "amgn_guide"],
        ["Base", "$22.50", "16x", "$360-365", "+11%", "amgn_guide"],
        ["Bull (obesity drug works)", "$23.50", "18x", "$423", "+30%", "amgn_guide"],
    ],
    "The simple P/E method ($360 base) agrees with the DCF ($388). Two independent methods pointing the same way (~$360-390 vs $326 today) raises confidence in the Buy.",
    widths={"A": 32, "B": 12, "C": 12, "D": 14, "E": 10, "F": 34},
)

add_sheet(
    "Amgen-Pipeline", "Amgen - drug pipeline & its value ($ billions of peak sales)",
    "Pipeline = drugs not yet selling. We estimate each one's peak annual sales and multiply by its probability of success (PoS) to 'risk-adjust' it. This is the upside not in today's revenue.",
    ["Drug", "What it treats", "Stage", "Launch", "Peak sales", "PoS", "Risk-adj.", "In base?", "Source"],
    [
        ["MariTide", "Obesity (monthly shot vs weekly rivals)", "Phase 3", "~2027-28", 6.0, "60%", 3.6, "No - upside", "maritide_ph2"],
        ["Olpasiran", "Heart risk (lowers Lp(a))", "Phase 3", "~2027-28", 2.5, "50%", 1.25, "No - upside", "pipeline"],
        ["Bemarituzumab", "Stomach cancer", "Phase 3", "~2027", 1.5, "55%", 0.8, "No - upside", "pipeline"],
        ["IMDELLTRA", "Small-cell lung cancer", "Approved+expanding", "selling", 3.0, "80%", 2.4, "Partly", "pipeline"],
        ["UPLIZNA", "Rare autoimmune diseases", "Approved+expanding", "selling", 2.0, "85%", 1.7, "Partly", "pipeline"],
        ["MariTide peak range (outside view)", "Evaluate $3.7bn (low) to UBS $10bn+ (high)", "-", "-", "-", "-", "-", "-", "maritide_pk"],
    ],
    "The new drugs add ~$5.6bn of risk-adjusted sales on top of today's ~$37bn from ~2028 - enough to restart growth after the Prolia patent loss. The wildcard is MariTide (obesity): outside estimates range from $3.7bn to $10bn+. You're buying that obesity 'lottery ticket' cheaply at today's price.",
    widths={"A": 26, "B": 40, "C": 18, "D": 11, "E": 11, "F": 7, "G": 10, "H": 11, "I": 32},
)

# ================= DANAHER =================
add_sheet(
    "Danaher-Segments", "Danaher - segment revenue build ($ billions)",
    "Danaher sells equipment/tools in three divisions. We forecast each one. Note the margins differ a lot, so the MIX (which division grows) matters as much as the total.",
    ["Segment", "FY2025", "Profit margin", "'26E gr.", "2026E", "'27E gr.", "2027E", "What drives it", "Source"],
    [
        ["Biotechnology (bioprocessing)", 7.3, "~38%", "+6%", 7.74, "+8%", 8.36, "Drug-makers re-ordering (orders +30%)", "dhr_seg"],
        ["Diagnostics (Cepheid, Beckman)", 10.0, "~31%", "+4%", 10.40, "+4%", 10.82, "Lab tests; China is a headwind", "dhr_seg"],
        ["Life Sciences (instruments/tools)", 7.3, "~21%", "+2%", 7.45, "+3%", 7.67, "Sluggish lab/academic spending", "dhr_seg"],
        ["ORGANIC TOTAL", 24.6, "~28%", "+4%", 25.6, "+5%", 26.9, "Matches 3-6% guidance", "dhr_overview"],
        ["+ Masimo (being acquired)", "-", "-", "-", 1.0, "-", 2.2, "Deal closes H2 2026", "dhr_tikr"],
        ["TOTAL REVENUE", 24.6, "", "", 26.6, "", 29.0, "", None],
    ],
    "The whole bull case is Biotechnology: it's the highest-margin division (~38%) AND the one re-accelerating (equipment orders +30%). As it grows, it lifts both company growth and profitability at once - that's what should push earnings (and the stock) higher.",
    widths={"A": 32, "B": 11, "C": 13, "D": 9, "E": 9, "F": 9, "G": 9, "H": 34, "I": 30},
    total_rows={3, 5},
)

add_sheet(
    "Danaher-Valuation", "Danaher - P/E valuation scenarios",
    "Earnings-per-share x P/E multiple, three scenarios. Danaher is a high-quality 'compounder' that historically trades at 25-30x; today it's ~22x (a discount).",
    ["Scenario", "EPS used", "P/E multiple", "Implied price", "vs ~$190", "Source"],
    [
        ["Bear (recovery stalls)", "2026E $8.45", "20x", "$169", "-11%", "dhr_tikr"],
        ["Base", "2027E ~$9.30", "24x", "$223", "+17%", "dhr_tikr"],
        ["Bull (full upcycle)", "2027E ~$9.60", "28x", "$269", "+42%", "dhr_tikr"],
    ],
    "You're buying a proven compounder at a below-average multiple right as its key division turns up. As earnings grow AND the multiple normalizes toward its history, you get a double benefit. Base target ~$223 (+17%).",
    widths={"A": 26, "B": 14, "C": 12, "D": 14, "E": 10, "F": 30},
)

# ================= SANDOZ =================
add_sheet(
    "Sandoz-Build", "Sandoz - revenue build: generics vs biosimilars ($ billions)",
    "Sandoz has two halves: declining/flat GENERICS and growing, higher-margin BIOSIMILARS. The investment question is whether biosimilars can grow fast enough to lift overall margins.",
    ["Segment", "FY2025", "'26E gr.", "2026E", "'27E gr.", "2027E", "Notes", "Source"],
    [
        ["Biosimilars", 3.29, "+14%", 3.75, "+15%", 4.31, "Copies of biologics; the growth engine (+15% in '25)", "sdz_fy25"],
        ["Generics", 7.79, "-1%", 7.71, "flat", 7.71, "Copies of pills; flat-to-down on pricing", "sdz_fy25"],
        ["TOTAL NET SALES", 11.09, "+5%", 11.46, "+5%", 12.02, "Mid-single-digit, matches guidance", "sdz_growth"],
        ["Biosimilar share of mix", "30%", "", "~33%", "", "~36%", "Rising mix = the margin lever", None],
    ],
    "The business works: biosimilars climb from 30% to ~36% of sales in two years, and because they're higher-margin, that shift funds the planned profit-margin improvement. The problem is purely VALUATION (next tab) - the good news is already in the price.",
    widths={"A": 22, "B": 10, "C": 9, "D": 10, "E": 9, "F": 10, "G": 40, "H": 34},
    total_rows={2},
)

add_sheet(
    "Sandoz-Valuation", "Sandoz - EV/EBITDA valuation scenarios",
    "Generics/biosimilars are valued on EV/EBITDA (enterprise value vs cash operating profit) rather than P/E. Three scenarios on ~$2.7bn of 2026E core EBITDA; net debt ~$3.6bn.",
    ["Scenario", "2026E core EBITDA", "EV/EBITDA", "Implied equity (less $3.6bn debt)", "Rough price", "Source"],
    [
        ["Bear (pricing worse)", "$2.6bn", "9x", "~$19.8bn", "~CHF 37", "sdz_growth"],
        ["Base", "$2.7bn", "11x", "~$26.1bn", "~CHF 48-50", "sdz_growth"],
        ["Bull (margin + biosimilar beats)", "$2.9bn", "13x", "~$34bn", "~CHF 63", "sdz_growth"],
    ],
    "Here's the catch: my BASE fair value (~CHF 48-50) is BELOW today's ~CHF 66 price. The market is already paying a premium for the biosimilar story. Great company, but no margin of safety - hence HOLD, not Buy. I'd get interested on a pullback toward the low-CHF-50s.",
    widths={"A": 30, "B": 17, "C": 11, "D": 32, "E": 14, "F": 30},
)

# ================= GUARDANT HEALTH =================
add_sheet(
    "GH-Financials", "Guardant Health - recent financials (actuals)",
    "Guardant sells blood tests for cancer. It is growing fast (~30-48%) but still LOSES money (burns cash) while it invests in screening. Segment splits are estimated; totals are reported.",
    ["Metric", "FY2024A", "FY2025A", "Q1 2026A", "Comment", "Source"],
    [
        ["Total revenue ($m)", 739.0, 982.0, 301.7, "FY25 +33%; Q1'26 +48% (accelerating)", "gh_fy25"],
        ["- Oncology (Guardant360 + Reveal)", 542.8, 683.6, "~205", "Therapy-selection tests; the cash engine", "gh_fy25"],
        ["- Screening (Shield)", "~25", 79.7, "~42", "Blood cancer-screening; ~44k tests Q1 vs ~9k", "gh_q1_bw"],
        ["- Biopharma & Data", 178.0, 210.1, "~53", "Selling data/testing to drug-makers", "gh_fy25"],
        ["Non-GAAP gross margin", "~63%", "~64%", "66%", "Screening margin jumped 18% -> 56%", "gh_q1_tr"],
        ["Adjusted EBITDA loss ($m)", -257.5, -220.9, "n/a", "Loss narrowing", "gh_fy25"],
        ["GAAP net loss ($m)", "n/a", -416.3, "n/a", "Inflated by stock comp + converts", "gh_fy25"],
        ["Free cash flow ($m)", -274.9, -233.1, "n/a", "Cash 'burn' - improving but still large", "gh_burn"],
        ["Cash & investments ($m)", "n/a", "n/a", "~1,200", "~4-5 yrs runway after Nov-25 raise", "gh_raise"],
    ],
    "Revenue is accelerating (+48% last quarter) and the economics of the Shield test improved sharply (gross margin 18% -> 56%). BUT the company still burns ~$200m+/yr and loses money. It has ~$1.2bn cash, so no crisis - but it's a growth bet, not a safe value stock.",
    widths={"A": 30, "B": 11, "C": 11, "D": 11, "E": 38, "F": 36},
)

add_sheet(
    "GH-RevenueBuild", "Guardant Health - revenue build by product ($ millions)",
    "We forecast each product line. Oncology = the steady cash engine; Shield = the screening rocket; Biopharma = a quiet annuity. Total checked vs guidance ($1.30-1.32bn for 2026).",
    ["Product line", "FY2024A", "FY2025A", "2026E", "2027E", "Key assumption", "Source"],
    [
        ["Oncology (Guardant360 + Reveal)", 542.8, 683.6, 875, 1085, ">35% volume / 28-29% revenue guide", "gh_guidance"],
        ["Screening (Shield)", 25, 79.7, 191, 334, "230-245k tests; guideline-driven ramp", "gh_guidance"],
        ["Biopharma & Data", 178, 210.1, 244, 278, "Steady mid-teens; Merck deals", "gh_fy25"],
        ["TOTAL REVENUE", 739, 982.0, 1310, 1697, "2026E matches guide $1.30-1.32bn", "gh_guidance"],
    ],
    "Oncology is still ~2/3 of revenue and pays the bills. But Shield (screening) is the slope that matters - it goes from 8% of sales (2025) to ~15% (2026E) to ~20% (2027E). If that keeps climbing, the stock re-rates higher. The entire long-term thesis is the next tab.",
    widths={"A": 34, "B": 11, "C": 11, "D": 10, "E": 10, "F": 38, "G": 32},
    total_rows={3},
)

add_sheet(
    "GH-Shield-TAM", "Guardant Health - Shield screening opportunity (TAM scenario)",
    "Shield is the whole investment case. This estimates its eventual (~2030) size: addressable people x how many get a blood test x Guardant's share x price per test. All figures illustrative ('E').",
    ["Driver", "Bear", "Base", "Bull", "Source"],
    [
        ["Addressable population, blood-amenable (millions)", 60, 80, 100, "shield_eclipse"],
        ["Annual % screened via blood test", "2%", "5%", "9%", None],
        ["Annual Shield tests in market (millions)", 1.2, 4.0, 9.0, None],
        ["Guardant's share of blood screening", "60%", "70%", "75%", "exact_free"],
        ["Net Shield tests to Guardant (millions)", 0.72, 2.8, 6.75, None],
        ["Blended price per test (ASP, $)", 700, 850, 1000, "shield_adlt"],
        ["IMPLIED SHIELD REVENUE ($bn)", "~0.5", "~2.4", "~6.8", None],
    ],
    "Even the BASE case (~$2.4bn) is ~2.5x Guardant's entire current revenue - from one product. That's why a seemingly expensive stock can still be cheap: you're paying for a franchise far bigger than today's P&L. The single biggest swing factor between Base and Bull is USPSTF approval (unlocks free insurance coverage).",
    widths={"A": 38, "B": 12, "C": 12, "D": 12, "E": 40},
    total_rows={6},
)

add_sheet(
    "GH-Competition", "Guardant Health - competition",
    "Who else is in cancer testing, and how big a threat they are. Guardant LEADS in blood-based colorectal screening but TRAILS Natera in recurrence testing (MRD).",
    ["Company / test", "Area", "Where it stands vs Guardant", "Threat", "Source"],
    [
        ["Exact Sciences (Cologuard) + Freenome", "CRC screening", "Incumbent ($3.2bn+ sales); blood test ~2 yrs behind Shield", "High (long-term)", "exact_guide"],
        ["Freenome (Exact-licensed)", "CRC screening (blood)", "~81% sensitivity; not yet FDA-approved", "High (via Exact)", "exact_free"],
        ["Natera - Signatera", "MRD / recurrence", "~$2.3bn sales, ~$28bn cap; the MRD leader", "High (vs Reveal)", "natera_10k"],
        ["Tempus AI", "Onc testing + data/AI", "MRD small base but growing fast", "Medium", "mrd_context"],
        ["Grail - Galleri", "Multi-cancer early detection", "Adjacent; not FDA-approved", "Medium", None],
        ["Roche (SAGA Diagnostics)", "MRD", "Deep pockets; longer-dated", "Medium", None],
    ],
    "Guardant is clearly ahead in blood CRC screening (only FDA primary-screen approval, locked Medicare price, guideline inclusion) with a ~2-year lead. But screening is a land-grab and a well-funded Exact/Freenome is coming ~late 2026. In recurrence testing (MRD), don't assume Guardant wins - Natera dominates, so Reveal is upside, not base.",
    widths={"A": 32, "B": 22, "C": 44, "D": 16, "E": 30},
)

add_sheet(
    "GH-Valuation", "Guardant Health - EV/Sales valuation (bear/base/bull)",
    "Guardant loses money, so we can't use P/E. Instead we value it on EV/Sales (enterprise value vs revenue), applied to 2027E revenue (~$1.7bn). It trades ~11x sales today.",
    ["Scenario", "2027E revenue", "EV/Sales", "Implied EV", "Per share", "vs ~$127", "Source"],
    [
        ["Bear (Shield stalls, multiple compresses)", "$1.55bn", "7x", "~$10.9bn", "~$82", "-35%", "gh_mult"],
        ["Base (Shield scales steadily)", "$1.70bn", "10x", "~$17.0bn", "~$150", "+18%", "gh_forecast"],
        ["Bull (Shield inflects + USPSTF)", "$1.85bn", "13x", "~$24.1bn", "~$181-205", "+43% to +61%", "btig"],
    ],
    "Wide outcomes (-35% to +61%) - that's the nature of a pre-profit growth stock. Base target ~$150 (+18%), skewed to upside because the huge Shield opportunity isn't in the financials yet. The check on the downside: ~$1.2bn cash = ~4-5 years runway, so it won't be forced to raise money cheaply soon. Size the position for volatility.",
    widths={"A": 36, "B": 13, "C": 10, "D": 12, "E": 12, "F": 12, "G": 30},
)

add_sheet(
    "GH-Catalysts", "Guardant Health - catalysts to watch (next 12 months)",
    "Events that could move the stock. The biggest by far is USPSTF (a screening recommendation that unlocks no-cost insurance coverage).",
    ["Catalyst", "Why it matters", "Rough timing", "Source"],
    [
        ["USPSTF recommendation for blood CRC screening", "Unlocks no-cost insurance coverage = volume surge", "2026-27", "shield_fda"],
        ["Quarterly Shield volume + price", "Proves the guideline win converts to real tests", "Every quarter", "gh_guidance"],
        ["Commercial-insurer coverage wins", "Broadens payment beyond Medicare's $1,495", "Ongoing", "shield_cov"],
        ["Exact/Freenome FDA filing & launch", "Competitive threat read-through", "~late 2026", "exact_free"],
        ["Reveal (MRD) reimbursement progress", "Upside optionality vs Natera", "2026", "natera_10k"],
        ["Cash-flow breakeven milestones", "Removes the dilution worry", "2027-28", "gh_burn"],
    ],
    "Watch USPSTF above all - it's the difference between the Base (~$2.4bn Shield) and Bull (~$6.8bn) cases. Also watch quarterly Shield volumes: that's the fastest proof the May-2026 guideline inclusion is turning into actual prescriptions.",
    widths={"A": 44, "B": 44, "C": 14, "D": 30},
)

# ================= SOURCES =================
ws = wb.create_sheet("Sources")
ws["A1"] = "Sources - every link used in this workbook"
ws["A1"].font = TITLE_FONT
ws.merge_cells("A1:C1")
ws["A2"] = "Click any link to open it. Note: SEC and some company/press pages block automated tools but open normally in a browser."
ws["A2"].font = INTRO_FONT
ws.merge_cells("A2:C2")
for c, h in enumerate(["#", "Source", "Link"], 1):
    cell = ws.cell(4, c, h); cell.fill = HDR_FILL; cell.font = HDR_FONT; cell.alignment = CENTER; cell.border = BORDER
for i, (key, (disp, url)) in enumerate(SRC.items(), 1):
    rr = 4 + i
    ws.cell(rr, 1, i).border = BORDER
    c2 = ws.cell(rr, 2, disp); c2.border = BORDER; c2.alignment = LEFT
    c3 = ws.cell(rr, 3, url); c3.hyperlink = url; c3.font = LINK_FONT; c3.border = BORDER; c3.alignment = LEFT
ws.column_dimensions["A"].width = 5
ws.column_dimensions["B"].width = 50
ws.column_dimensions["C"].width = 95
ws.freeze_panes = "A5"

wb.save("/home/user/aanyapatel3.github.io/EquityResearch_Master.xlsx")
print("Wrote EquityResearch_Master.xlsx with", len(wb.sheetnames), "sheets:")
for s in wb.sheetnames:
    print("  -", s)
